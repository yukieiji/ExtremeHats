import os
import json
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))

HAT_IN_FILE = os.path.join(WORKING_DIR, 'HatTransData.xlsx')
HAT_OUT_FILE = os.path.join(WORKING_DIR, 'hat', 'hatTranData.json')
HAT_FOLDER = os.path.join(WORKING_DIR, 'hat')
HAT_DATA_FILE = os.path.join(WORKING_DIR, 'hat', 'hatData.json')

HAT_DATA_DATA_KEY = 'data'
HAT_DATA_UPDATE_COMIT_KEY = 'updateComitHash'

def string_to_json(filename, outputFile) -> None:
  wb = load_workbook(filename, read_only = True)

  stringData = {}
  for s in wb:
    rows = s.iter_rows(min_col = 1, min_row = 2, max_col = 17, max_row = None)
    headers = []
    for header in s[1]:
      if header.value:
        headers.append(header.value)

    for row in rows:
      name = row[0].value

      if not name:
        continue

      data = {}

      for i, string in enumerate(row[1:]):
        if string.value:
          # I hate excel why did I do this to myself
          data[i] = string.value.replace("\r", "").replace("_x000D_", "").replace("\\n", "\n")

      if data:
        stringData[name] = data

  with open(outputFile, "w") as f:
    json.dump(stringData, f, indent=4)

def update_hat_data():

  hats = os.listdir(HAT_FOLDER)

  with open(os.path.join(HAT_DATA_FILE), mode='w') as hat_json:
    json.dump(
      {HAT_DATA_DATA_KEY:hats, HAT_DATA_UPDATE_COMIT_KEY:""},
      hat_json, indent=2, ensure_ascii=False)

def main():
  string_to_json(HAT_IN_FILE, HAT_OUT_FILE)
  update_hat_data()

if __name__ == "__main__":
  main()